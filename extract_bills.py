#!/usr/bin/env python3
"""
extract_bills.py
Extrae metadata estructurada de los PDFs de proyectos de ley
y genera bills_data.json para el dashboard.
"""

import pdfplumber
import json
import os
import re
import unicodedata
from datetime import datetime
from collections import Counter
from openpyxl import load_workbook

# ─── DICCIONARIO DE TEMAS (palabras clave → tema) ──────────────────────────
TOPICS = {
    "Tecnología / IA": [
        "inteligencia artificial", "ia ", " ia,", "digital", "tecnolog",
        "innovación", "ciberseguridad", "datos", "software", "algoritmo",
        "internet", "nodo", "neurotecnolog", "neurodato"
    ],
    "Educación": [
        "educaci", "escuela", "docente", "universid", "enseñanza",
        "estudiante", "aula", "pedagog", "formaci", "becas", "conicet",
        "ciencia", "investigaci"
    ],
    "Salud": [
        "salud", "médic", "hospital", "medicamento", "sanitari",
        "enfermedad", "paciente", "farmac", "clínica", "epidemi"
    ],
    "Economía / Finanzas": [
        "econom", "fiscal", "presupuest", "impuest", "financi",
        "fondo", "banco", "inversión", "crédito", "deuda", "mercado",
        "commercial", "comerci"
    ],
    "Ambiente": [
        "ambient", "ecolog", "climat", "cambio climático", "residuo",
        "glaciar", "minería", "biodiversidad", "contaminaci", "sustentabl",
        "renovable", "energía"
    ],
    "Género / Diversidad": [
        "género", "femini", "mujer", "violencia de género", "diversidad",
        "lgb", "paridad", "discriminaci", "identidad"
    ],
    "Seguridad / Justicia": [
        "seguridad", "justicia", "penal", "delito", "policía",
        "crimen", "prisión", "narcotráf", "corrupci", "fuero"
    ],
    "Trabajo / Social": [
        "trabajo", "laboral", "empleo", "sindical", "jubilación",
        "pensión", "social", "pobreza", "vivienda", "habitat"
    ],
    "Infraestructura": [
        "infraestructura", "obra", "vial", "transport", "ferrocarril",
        "aeropuerto", "puerto", "autopista", "camino"
    ],
    "Soberanía / Defensa": [
        "soberanía", "defensa", "estratégic", "activo estratégico",
        "nacional", "fuerzas armadas", "hidrocarburos", "recurso natural"
    ],
}

# ─── BLOQUE LEGISLATIVO (curado para los autores de esta base) ──────────────
# Fuentes: perfiles/proyectos oficiales HCDN + firma en texto del propio TP.
BLOQUE_MAP = {
    # La Libertad Avanza
    "GABRIEL M BORNORONI": "La Libertad Avanza",
    "BORNORONI GABRIEL M": "La Libertad Avanza",
    "SILVANA GIUDICI": "La Libertad Avanza",
    "MARCELA MARINA PAGANO": "La Libertad Avanza",
    "PAGANO MARCELA MARINA": "La Libertad Avanza",

    # PRO
    "GERARDO MILMAN": "PRO",
    "VICTORIA MORALES GORLERI": "PRO",
    "MORALES GORLERI VICTORIA": "PRO",

    # Unión por la Patria / Frente de Todos
    "DANTE LOPEZ RODRIGUEZ": "Unión por la Patria",
    "LOPEZ RODRIGUEZ DANTE": "Unión por la Patria",
    "DIEGO A GIULIANO": "Unión por la Patria",
    "GIULIANO DIEGO A": "Unión por la Patria",
    "DANIEL GOLLAN": "Unión por la Patria",
    "GOLLAN DANIEL": "Unión por la Patria",
    "GISELA MARZIOTTA": "Unión por la Patria",
    "MARZIOTTA GISELA": "Unión por la Patria",
    "SILVANA MICAELA GINOCCHIO": "Unión por la Patria",
    "GINOCCHIO SILVANA MICAELA": "Unión por la Patria",
    "ANAHI COSTA": "Frente de Todos",
    "COSTA ANAHI": "Frente de Todos",

    # Espacios federales
    "PABLO OUTES": "Innovación Federal",
    "OUTES PABLO": "Innovación Federal",
    "YOLANDA VEGA": "Innovación Federal",
    "VEGA YOLANDA": "Innovación Federal",
    "PAMELA CALLETTI": "Innovación Federal",
    "CALLETTI PAMELA": "Innovación Federal",
    "JUAN FERNANDO BRUGGE": "Encuentro Federal",
    "BRUGGE JUAN FERNANDO": "Encuentro Federal",
    "OSCAR AGOST CARRENO": "Encuentro Federal",
    "AGOST CARRENO OSCAR": "Encuentro Federal",
    "MONICA FEIN": "Encuentro Federal",

    # UCR / aliados
    "JIMENA LATORRE": "UCR",
    "LATORRE JIMENA": "UCR",
    "NATALIA SILVINA SARAPURA": "UCR",
    "SARAPURA NATALIA SILVINA": "UCR",
    "MAXIMILIANO FERRARO": "Coalición Cívica",

    # Democracia para Siempre
    "DANYA TAVELA": "Democracia para Siempre",
    "TAVELA DANYA": "Democracia para Siempre",
    "MANUEL IGNACIO AGUIRRE": "Democracia para Siempre",
    "AGUIRRE MANUEL IGNACIO": "Democracia para Siempre",

    # Otros bloques relevantes del período
    "GABRIEL FELIPE CHUMPITAZ": "Futuro y Libertad",
    "CHUMPITAZ GABRIEL FELIPE": "Futuro y Libertad",
    "ROSSO VICTORIA": "Unidad Justicialista",
    "VICTORIA ROSSO": "Unidad Justicialista",
}

# ─── DESTINATARIOS (palabras clave → destinatario) ────────────────────────
DESTINATARIOS_MAP = {
    "Ciudadanía general": [
        "persona", "ciudadan", "habitante", "individuo", "poblaci",
        "usuario", "sociedad", "público en general",
    ],
    "Investigadores / Científicos": [
        "investigador", "científico", "becario", "conicet", "investigaci",
        "académic", "universid", "posgrado",
    ],
    "Sector educativo": [
        "docente", "alumno", "estudiante", "escuela", "educaci",
        "formaci", "enseñanza", "pedagog",
    ],
    "Empresas / Sector privado": [
        "empresa", "emprendedor", "pyme", "industria", "sector productivo",
        "comercio", "inversión", "negocio", "startup",
    ],
    "Sector tecnológico / IA": [
        "proveedor de ia", "desarrollador", "tecnolog", "plataforma digital",
        "operador de algoritmo", "sistema de inteligencia",
    ],
    "Salud": [
        "paciente", "médic", "hospital", "sanitari", "clínic",
        "farmac", "profesional de salud",
    ],
    "Estado / Organismo público": [
        "organism", "adminis", "ministerio", "gobierno", "estado nacional",
        "poder ejecutiv", "poder legislativ", "poder judicial",
    ],
    "Jóvenes / Estudiantes": [
        "jóvenes", "joven", "menor", "niño", "adolescente",
        "universitario", "estudiante universitario",
    ],
}

PDF_DIRS = [".", "Mateo proyectos"]


EXCLUDED_BILL_IDS = {"0345-D-2025"}


XLSX_CANDIDATES = ["Proyectos de IA (1).xlsx", "Proyectos de IA.xlsx"]


def normalize_bill_id(raw_id: str) -> str:
    if not raw_id:
        return ""
    m = re.match(r"^(\d{3,4})-?([DS])-(\d{4})$", str(raw_id).strip().upper().replace(" ", ""))
    if not m:
        return ""
    num, cam, year = m.groups()
    return f"{int(num):04d}-{cam}-{year}"


def load_xlsx_curated(script_dir: str) -> dict:
    """Carga columnas curadas (tipo/dashboard/bloque/autor) desde el XLSX de proyectos."""
    curated = {}
    xlsx_path = None
    for name in XLSX_CANDIDATES:
        cand = os.path.join(script_dir, name)
        if os.path.exists(cand):
            xlsx_path = cand
            break
    if not xlsx_path:
        return curated

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"⚠️ No se pudo abrir XLSX curado: {e}")
        return curated

    if "PROYECTOS" not in wb.sheetnames:
        return curated

    ws = wb["PROYECTOS"]
    headers = [str(ws.cell(1, c).value).strip().upper() if ws.cell(1, c).value else ""
               for c in range(1, ws.max_column + 1)]

    def col(name):
        return headers.index(name) + 1 if name in headers else None

    c_num = col("NUMERO")
    c_tipo = col("TIPO")
    c_dash = col("DASHBOARD")
    c_bloque = col("BLOQUE")
    c_autor = col("AUTOR")

    if not c_num:
        return curated

    for r in range(2, ws.max_row + 1):
        bid = normalize_bill_id(ws.cell(r, c_num).value)
        if not bid:
            continue

        tipo = ws.cell(r, c_tipo).value if c_tipo else None
        dash = ws.cell(r, c_dash).value if c_dash else None
        bloque = ws.cell(r, c_bloque).value if c_bloque else None
        autor = ws.cell(r, c_autor).value if c_autor else None

        row = {}
        if tipo not in (None, ""):
            row["tipo"] = str(tipo).strip()
        if dash not in (None, ""):
            row["dashboard"] = str(dash).strip().upper()
        if bloque not in (None, ""):
            row["bloque_xlsx"] = str(bloque).strip()
        if autor not in (None, ""):
            row["autor_xlsx"] = str(autor).strip()

        if row:
            curated[bid] = row

    return curated


def infer_topic(text: str) -> str:
    """Infiere el tema principal del proyecto por palabras clave."""
    text_lower = text.lower()
    scores = {}
    for topic, keywords in TOPICS.items():
        score = sum(1 for kw in keywords if kw in text_lower)
        if score > 0:
            scores[topic] = score
    if scores:
        return max(scores, key=scores.get)
    return "Otros"


def infer_destinatarios(text: str) -> list:
    """Infiere los destinatarios principales del proyecto."""
    text_lower = text.lower()
    found = []
    for dest, keywords in DESTINATARIOS_MAP.items():
        score = sum(1 for kw in keywords if kw in text_lower)
        if score >= 2:  # Al menos 2 menciones
            found.append((dest, score))
    found.sort(key=lambda x: x[1], reverse=True)
    result = [d[0] for d in found[:3]]  # Up to 3 destinatarios
    return result if result else ["Ciudadanía general"]


def normalize_name(text: str) -> str:
    """Normaliza nombres para matching robusto (acentos/puntuación/espacios)."""
    if not text:
        return ""
    t = unicodedata.normalize("NFKD", text)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.upper()
    t = re.sub(r"[^A-Z0-9\s]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def infer_bloque(autor: str) -> str:
    """Devuelve bloque legislativo del autor según mapeo curado."""
    autor_norm = normalize_name(autor)
    for alias, bloque in BLOQUE_MAP.items():
        if alias in autor_norm:
            return bloque
    return "Sin datos"


def canonicalize_bloque(bloque: str) -> str:
    """Normaliza etiquetas de bloque para evitar duplicados por formato."""
    if not bloque:
        return "Sin datos"
    raw = str(bloque).strip()
    norm = normalize_name(raw)
    canon = {
        "PRO": "PRO",
        "UCR": "UCR",
        "COALICION CIVICA": "Coalición Cívica",
        "LA LIBERTAD AVANZA": "La Libertad Avanza",
        "LLA": "La Libertad Avanza",
        "UNION POR LA PATRIA": "Unión por la Patria",
        "FRENTE DE TODOS": "Frente de Todos",
        "INNOVACION FEDERAL": "Innovación Federal",
        "ENCUENTRO FEDERAL": "Encuentro Federal",
        "PROVINCIAS UNIDAS": "Provincias Unidas",
        "DEMOCRACIA PARA SIEMPRE": "Democracia para Siempre",
        "FUTURO Y LIBERTAD": "Futuro y Libertad",
        "UNIDAD JUSTICIALISTA": "Unidad Justicialista",
        "COHERENCIA": "COHERENCIA",
        "SIN DATOS": "Sin datos",
    }
    return canon.get(norm, raw)


def extract_title(lines: list) -> str:
    """Extrae el título del proyecto (líneas en MAYÚSCULAS luego del encabezado)."""
    skip_patterns = [
        r'^\d{4}\s*-\s*año', r'^proyecto de ley', r'^el senado',
        r'^la cámara', r'^sancionan', r'^"', r'^art[íi]culo',
        r'^\d+$', r'^nación', r'^www\.'
    ]
    title_parts = []
    found_title = False
    for line in lines[:30]:
        line_clean = line.strip()
        if not line_clean:
            if found_title and title_parts:
                break
            continue
        skip = any(re.match(p, line_clean.lower()) for p in skip_patterns)
        if skip:
            continue
        # El título suele estar en MAYÚSCULAS
        if line_clean == line_clean.upper() and len(line_clean) > 5:
            title_parts.append(line_clean)
            found_title = True
        elif found_title:
            # Primera línea no-mayúsculas después del título → parar
            break

    if title_parts:
        return " ".join(title_parts).title()
    # Fallback: primera línea sustancial
    for line in lines[:20]:
        if len(line.strip()) > 10:
            return line.strip().title()
    return "Sin título"


def extract_authors(text: str, lines: list) -> list:
    """Extrae autores del final del documento."""
    authors = []

    # Patrón 1: NOMBRE APELLIDO seguido de DIPUTAD/SENADOR
    pattern1 = re.findall(
        r'([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑA-Z\s\-]{3,40})\s*\n\s*(?:DIPUTAD[OA]|SENADOR[A])',
        text, re.IGNORECASE
    )
    for a in pattern1:
        clean = a.strip().title()
        if len(clean) > 3 and clean not in authors:
            authors.append(clean)

    # Patrón 2: FIRMANTE: / ACOMPAÑA: líneas
    for line in lines:
        m = re.match(r'(?:firmante|acompaña)\s*:\s*diputad[oa]\s+(.+)', line.strip(), re.IGNORECASE)
        if m:
            names = [n.strip().title() for n in m.group(1).split('-') if n.strip()]
            for n in names:
                if n not in authors:
                    authors.append(n)

    # Patrón 3: últimas líneas con nombres completos (antes de "DIPUTADO/A NACIONAL")
    last_lines = lines[-10:]
    for i, line in enumerate(last_lines):
        if re.search(r'diputad[oa]\s+nacional|senador[a]\s+nacional', line, re.IGNORECASE):
            if i > 0:
                candidate = last_lines[i-1].strip().title()
                if (len(candidate) > 4 and
                        not re.search(r'proyecto|ley|art[íi]culo|solicito', candidate, re.IGNORECASE) and
                        candidate not in authors):
                    authors.append(candidate)
        # Multiple authors separated by -
        elif re.match(r'^[A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s\-]+$', line.strip()) and len(line.strip()) > 6:
            # Could be a line with multiple authors
            pass

    # Limpiar autores vacíos o muy cortos
    authors = [a for a in authors if len(a) > 3]

    if not authors:
        # Último recurso: líneas en mayúsculas al final
        for line in reversed(lines[-15:]):
            l = line.strip()
            if (l == l.upper() and len(l) > 4 and
                    not re.search(r'ARTÍCULO|LEY|PROYECTO|NACIÓN|ARGENTINA', l)):
                # Split by -
                parts = [p.strip().title() for p in l.split('-') if p.strip() and len(p.strip()) > 3]
                if parts:
                    authors = parts
                    break

    return authors if authors else ["Desconocido"]


def infer_month_from_id(bill_id: str) -> int:
    """Infiere el mes aproximado según el número de expediente."""
    try:
        num = int(bill_id.split('-')[0])
        # En Diputados los números van ~0001 (enero) a ~8000+ (diciembre)
        return max(1, min(12, int((num / 800) + 1)))
    except:
        return 1


def bill_sort_key(bill_id: str) -> tuple:
    """Orden cronológico estable por año y número de expediente."""
    parts = bill_id.split("-")
    try:
        num = int(parts[0])
    except Exception:
        num = 999999
    try:
        year = int(parts[2])
    except Exception:
        year = 0
    camara = parts[1] if len(parts) > 1 else "D"
    return (year, num, camara, bill_id)


def discover_pdf_paths(script_dir: str) -> dict:
    """Descubre PDFs en múltiples carpetas y deduplica por ID."""
    by_id = {}
    for rel_dir in PDF_DIRS:
        abs_dir = os.path.join(script_dir, rel_dir)
        if not os.path.isdir(abs_dir):
            continue
        for filename in sorted(os.listdir(abs_dir)):
            if not filename.lower().endswith(".pdf"):
                continue
            bill_id = filename[:-4]
            if bill_id in by_id:
                continue
            rel_path = os.path.relpath(os.path.join(abs_dir, filename), script_dir).replace("\\", "/")
            by_id[bill_id] = rel_path
    return by_id


def load_existing_bills(output_path: str) -> dict:
    """Carga bills_data.json previo para preservar campos curados manualmente."""
    if not os.path.exists(output_path):
        return {}
    try:
        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {b["id"]: b for b in data.get("bills", []) if b.get("id")}
    except Exception as e:
        print(f"⚠️ No se pudo leer {output_path}: {e}")
        return {}


def merge_with_existing(parsed_bill: dict, existing_bill: dict, rel_pdf_path: str, curated_row: dict = None) -> dict:
    """Mezcla dato parseado con dato existente, priorizando campos manuales existentes."""
    merged = dict(parsed_bill)

    if existing_bill:
        recomputed_fields = {"bloque"}
        for k, v in existing_bill.items():
            if k in {"id", "pdf_path"}:
                continue
            if k in recomputed_fields:
                continue
            if v not in (None, "", []):
                merged[k] = v

    # Si el parseo del PDF no detecta autor correctamente,
    # recuperamos bloque desde autor existente previo.
    if merged.get("bloque") == "Sin datos":
        fallback_author = ""
        if existing_bill and existing_bill.get("autor_principal"):
            fallback_author = existing_bill.get("autor_principal", "")
        elif merged.get("autor_principal"):
            fallback_author = merged.get("autor_principal", "")
        if fallback_author:
            inferred = infer_bloque(fallback_author)
            if inferred != "Sin datos":
                merged["bloque"] = inferred

    # Enriquecimiento por XLSX curado (prioridad para completar faltantes)
    curated_row = curated_row or {}
    if curated_row.get("tipo") and not merged.get("tipo"):
        merged["tipo"] = curated_row["tipo"]
    if curated_row.get("dashboard") and not merged.get("dashboard"):
        merged["dashboard"] = curated_row["dashboard"]
    if curated_row.get("bloque_xlsx"):
        merged["bloque"] = curated_row["bloque_xlsx"]

    # Canonicalizar bloque para mantener una sola etiqueta por espacio político
    merged["bloque"] = canonicalize_bloque(merged.get("bloque", "Sin datos"))

    # Normalizar defaults requeridos por validación
    if not merged.get("tipo"):
        merged["tipo"] = "SIN CLASIFICAR"
    if not merged.get("dashboard"):
        merged["dashboard"] = "INDIRECTA"

    merged["pdf_path"] = f"./{rel_pdf_path}"
    return merged


def parse_pdf(filepath: str) -> dict:
    """Extrae datos estructurados de un PDF."""
    filename = os.path.basename(filepath)
    # Parse filename: XXXX-C-YYYY.pdf
    parts = filename.replace('.pdf', '').split('-')
    bill_num = parts[0] if parts else "0000"
    camara_code = parts[1] if len(parts) > 1 else "D"
    year = int(parts[2]) if len(parts) > 2 else 2025

    camara = "Diputados" if camara_code == "D" else "Senado"

    all_text = ""
    lines = []
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    all_text += t + "\n"
        lines = [l.strip() for l in all_text.split('\n') if l.strip()]
    except Exception as e:
        print(f"  ERROR leyendo {filename}: {e}")
        return None

    title = extract_title(lines)
    authors = extract_authors(all_text, lines)
    topic = infer_topic(all_text[:3000])  # Solo los primeros 3000 chars
    month = infer_month_from_id(bill_num)

    # Resumen: primera oración del artículo 1
    resumen = ""
    for i, line in enumerate(lines):
        if re.match(r'art[íi]culo\s*1', line, re.IGNORECASE):
            resumen_parts = []
            for l in lines[i:i+5]:
                resumen_parts.append(l)
                if len(" ".join(resumen_parts)) > 150:
                    break
            resumen = " ".join(resumen_parts)[:200] + "..."
            break

    bloque = infer_bloque(authors[0] if authors else "")
    destinatarios = infer_destinatarios(all_text)

    return {
        "id": filename.replace('.pdf', ''),
        "titulo": title,
        "autores": authors,
        "autor_principal": authors[0] if authors else "Desconocido",
        "total_autores": len(authors),
        "bloque": bloque,
        "camara": camara,
        "año": year,
        "mes": month,
        "tema": topic,
        "estado": "En Comisión",
        "destinatarios": destinatarios,
        "resumen": resumen,
        "expediente": f"{bill_num}/{year}",
    }


# ─── DATOS HISTÓRICOS SIMULADOS (2015–2024) ──────────────────────────────────

LEGISLADORES_SIMULADOS = [
    "Mario Negri", "Elisa Carrió", "Sergio Massa", "Axel Kicillof",
    "Martín Lousteau", "Graciela Camaño", "Victoria Villarruel",
    "Luis Petri", "Emilio Monzó", "Paula Oliveto Lago",
    "Juan Manuel López", "Marcela Campagnoli", "Maximiliano Ferraro",
    "Esteban Paulón", "Mónica Fein", "Pamela Calletti", "Diego Bossio",
    "Cristian Ritondo", "Waldo Wolff", "Silvia Lospennato",
    "Fernando Iglesias", "Pablo Tonelli", "Rogelio Frigerio",
    "Alfredo Cornejo", "Eduardo Bucca", "Graciela Ocaña",
]

TITULOS_SIMULADOS = {
    "Tecnología / IA": [
        "Régimen de Promoción de la Economía del Conocimiento",
        "Marco Regulatorio de Inteligencia Artificial",
        "Ley de Gobierno Digital y Servicios Electrónicos",
        "Programa Nacional de Ciberseguridad",
        "Regulación del Trabajo en Plataformas Digitales",
    ],
    "Educación": [
        "Ley de Financiamiento Educativo",
        "Creación del Programa Nacional de Alfabetización Digital",
        "Modificación de la Ley de Educación Superior",
        "Fondo de Becas para Estudiantes de Bajos Recursos",
        "Régimen de la Educación Técnico Profesional",
    ],
    "Salud": [
        "Ley de Acceso a Medicamentos Genéricos",
        "Creación del Sistema Nacional de Salud Mental",
        "Regulación de la Telemedicina",
        "Ley de Trasplante de Órganos — Donación Presunta",
        "Programa de Prevención de Enfermedades Crónicas",
    ],
    "Economía / Finanzas": [
        "Reforma del Sistema Tributario Nacional",
        "Ley de Inversiones Extranjeras Directas",
        "Creación del Fondo Soberano de Estabilización",
        "Régimen de Deuda Pública y Transparencia Fiscal",
        "Ley de PyMEs y Emprendedores",
    ],
    "Ambiente": [
        "Ley de Presupuestos Mínimos para el Cambio Climático",
        "Régimen de Energías Renovables",
        "Protección de Glaciares y Periglacial",
        "Gestión Integral de Residuos Sólidos Urbanos",
        "Ley de Humedales",
    ],
    "Género / Diversidad": [
        "Ley Micaela — Capacitación en Género",
        "Cupo Laboral Travesti-Trans",
        "Ampliación de la Ley de Paridad de Género",
        "Violencia Económica y Patrimonial",
        "Plan Nacional de Prevención del Femicidio",
    ],
    "Seguridad / Justicia": [
        "Reforma del Código Procesal Penal",
        "Creación de la Unidad Fiscal de Crimen Organizado",
        "Ley de Narcomenudeo y Prevención",
        "Regulación del Sistema Penitenciario Federal",
        "Modificación del Régimen de Menores en Conflicto con la Ley",
    ],
    "Trabajo / Social": [
        "Ley de Teletrabajo — Actualización",
        "Creación del Seguro de Desempleo Universal",
        "Reforma del Sistema Jubilatorio",
        "Ley de Economía Social y Solidaria",
        "Programa Nacional de Vivienda Popular",
    ],
    "Infraestructura": [
        "Plan Nacional de Obras Hidráulicas",
        "Concesión de Autopistas y Rutas Nacionales",
        "Modernización del Ferrocarril Belgrano Cargas",
        "Ley de Obras Públicas — Transparencia y Control",
        "Proyecto de Ampliación Aeropuerto Internacional Ezeiza",
    ],
    "Soberanía / Defensa": [
        "Ley de Defensa Nacional — Actualización",
        "Régimen de Activos Estratégicos Nacionales",
        "Proyecto de Modernización de las Fuerzas Armadas",
        "Soberanía Tecnológica — Protección de Infraestructura Crítica",
        "Ley de Inteligencia Nacional",
    ],
    "Otros": [
        "Modificación de Diversas Normas Legales",
        "Creación de Comisión Bicameral de Seguimiento",
        "Declaración de Interés Nacional",
        "Restitución de Vigencia de Artículos Derogados",
        "Fondo Especial para Zonas de Emergencia",
    ],
}

YEARS_COUNTS = {
    2015: 7842, 2016: 8123, 2017: 8567, 2018: 7934, 2019: 8901,
    2020: 5234, 2021: 7123, 2022: 8456, 2023: 9012, 2024: 8234,
}

ESTADOS_DIST = {
    "En Comisión": 0.62,
    "Aprobado": 0.24,
    "Archivado": 0.14,
}

MONTHLY_WEIGHTS = [0.06, 0.07, 0.09, 0.09, 0.10, 0.09, 0.08, 0.07, 0.10, 0.10, 0.09, 0.06]


def generate_historical_data() -> list:
    """Genera datos históricos simulados para 2015–2024."""
    random.seed(42)
    bills = []
    bill_id_counter = 1

    topics_list = list(TOPICS.keys()) + ["Otros"]
    topic_weights = [0.18, 0.15, 0.10, 0.12, 0.08, 0.07, 0.07, 0.08, 0.06, 0.05, 0.04]

    for year, total in YEARS_COUNTS.items():
        # Generar una muestra representativa (no todos los proyectos, sino ~200/año para el demo)
        sample_size = 200
        for _ in range(sample_size):
            topic = random.choices(topics_list, weights=topic_weights)[0]
            titulo_list = TITULOS_SIMULADOS.get(topic, TITULOS_SIMULADOS["Otros"])
            titulo = random.choice(titulo_list)
            autor = random.choice(LEGISLADORES_SIMULADOS)
            camara = random.choices(CAMARAS, weights=[0.6, 0.4])[0]
            month = random.choices(range(1, 13), weights=MONTHLY_WEIGHTS)[0]
            # Estado ponderado
            estado = random.choices(
                list(ESTADOS_DIST.keys()),
                weights=list(ESTADOS_DIST.values())
            )[0]

            bills.append({
                "id": f"{bill_id_counter:04d}-{'D' if camara == 'Diputados' else 'S'}-{year}",
                "titulo": titulo,
                "autores": [autor],
                "autor_principal": autor,
                "camara": camara,
                "año": year,
                "mes": month,
                "tema": topic,
                "estado": estado,
                "resumen": f"Proyecto presentado en {year} sobre {topic.lower()}.",
                "expediente": f"{bill_id_counter:04d}/{year}",
                "simulado": True,
            })
            bill_id_counter += 1

    return bills


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "bills_data.json")
    existing_by_id = load_existing_bills(output_path)
    curated_by_id = load_xlsx_curated(script_dir)
    pdf_map = discover_pdf_paths(script_dir)
    pdf_ids = sorted(pdf_map.keys(), key=bill_sort_key)
    if EXCLUDED_BILL_IDS:
        pdf_ids = [bid for bid in pdf_ids if bid not in EXCLUDED_BILL_IDS]

    print(f"Encontrados {len(pdf_ids)} PDFs únicos en {', '.join(PDF_DIRS)}")

    bills = []
    for bill_id in pdf_ids:
        rel_path = pdf_map[bill_id]
        filepath = os.path.join(script_dir, rel_path)
        print(f"  Procesando {rel_path}...", end=" ")
        parsed_bill = parse_pdf(filepath)
        existing_bill = existing_by_id.get(bill_id)

        if parsed_bill:
            bill = merge_with_existing(parsed_bill, existing_bill, rel_path, curated_by_id.get(bill_id))
            bills.append(bill)
            print(f"✓ '{bill['titulo'][:45]}' | {bill['tema']} | {bill['bloque']}")
        elif existing_bill:
            fallback = dict(existing_bill)
            fallback["pdf_path"] = f"./{rel_path}"
            bills.append(fallback)
            print("⚠️ Se preserva versión previa por error de lectura")
        else:
            print("✗ Error")

    # Bloque stats
    bloque_counter = Counter(b["bloque"] for b in bills)
    bloque_autores = {}
    for b in bills:
        bl = b["bloque"]
        bloque_autores.setdefault(bl, {"proyectos": 0, "autores_total": 0})
        bloque_autores[bl]["proyectos"] += 1
        bloque_autores[bl]["autores_total"] += b["total_autores"]

    bloque_stats = [
        {"bloque": bl, "proyectos": v["proyectos"],
         "promedio_autores": round(v["autores_total"] / v["proyectos"], 1)}
        for bl, v in bloque_autores.items()
    ]
    bloque_stats.sort(key=lambda x: x["proyectos"], reverse=True)

    # Destinatarios stats
    dest_counter = Counter()
    for b in bills:
        for d in b["destinatarios"]:
            dest_counter[d] += 1
    top_destinatarios = [{"destinatario": d, "total": c}
                         for d, c in dest_counter.most_common()]

    # Topic stats
    topic_counter = Counter(b["tema"] for b in bills)
    top_topics = [{"tema": t, "total": c} for t, c in topic_counter.most_common()]

    output = {
        "metadata": {
            "generado": datetime.now().isoformat(),
            "total": len(bills),
            "fuente": "PDFs locales del proyecto (raíz + Mateo proyectos)",
            "carpetas_fuente": PDF_DIRS,
        },
        "bloque_stats": bloque_stats,
        "top_destinatarios": top_destinatarios,
        "top_topics": top_topics,
        "bills": bills,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Exportado: {output_path}")
    print(f"   {len(bills)} proyectos reales")
    print(f"   Bloques: {dict(bloque_counter)}")
    print(f"   Top destinatarios: {[d['destinatario'] for d in top_destinatarios[:3]]}")


if __name__ == "__main__":
    main()
